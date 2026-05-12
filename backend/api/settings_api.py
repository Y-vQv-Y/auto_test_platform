"""系统设置 API - 安全、登录信息、仪表盘等"""
import os
from fastapi import APIRouter, Depends, Query
from sqlalchemy.orm import Session
from pydantic import BaseModel
from typing import Optional, List
from datetime import datetime, timedelta, timezone
from loguru import logger

from backend.database import get_db, Project, TestRun, TestCase, AIConfig, SecurityLog, LoginRecord
from backend.config import settings
from backend.security import URLValidator, ToolWhitelist
from backend.captcha import CaptchaHandler
from backend.errors import ErrorCode, error_response

router = APIRouter(prefix="/api/v1", tags=["系统设置"])


# ========== 仪表盘 ==========
@router.get("/dashboard")
def get_dashboard(db: Session = Depends(get_db)):
    """获取仪表盘数据"""
    now = datetime.now(timezone.utc)
    week_ago = now - timedelta(days=7)

    # 统计数据
    total_projects = db.query(Project).filter(Project.status == "active").count()
    total_test_runs = db.query(TestRun).count()
    total_test_cases = db.query(TestCase).count()
    total_ai_configs = db.query(AIConfig).filter(AIConfig.status == "active").count()

    # 最近运行
    recent_runs = (
        db.query(TestRun)
        .order_by(TestRun.created_at.desc())
        .limit(10)
        .all()
    )

    # 本周运行统计
    weekly_runs = db.query(TestRun).filter(
        TestRun.created_at >= week_ago
    ).count()
    weekly_passed = db.query(TestRun).filter(
        TestRun.created_at >= week_ago,
        TestRun.status == "passed",
    ).count()

    # 成功率（计算所有已完成的运行：passed + failed + error）
    success_rate = 0
    completed_runs = db.query(TestRun).filter(
        TestRun.status.in_(["passed", "failed", "error"])
    )
    total_completed = completed_runs.count()
    if total_completed > 0:
        passed_count = completed_runs.filter(TestRun.status == "passed").count()
        success_rate = round(passed_count / total_completed * 100, 1)

    # 状态分布
    status_distribution = {}
    for status in ["pending", "running", "passed", "failed", "error"]:
        count = db.query(TestRun).filter(TestRun.status == status).count()
        if count > 0:
            status_distribution[status] = count

    return {
        "stats": {
            "total_projects": total_projects,
            "total_test_runs": total_test_runs,
            "total_test_cases": total_test_cases,
            "total_ai_configs": total_ai_configs,
            "weekly_runs": weekly_runs,
            "weekly_passed": weekly_passed,
            "success_rate": success_rate,
        },
        "status_distribution": status_distribution,
        "recent_runs": [
            {
                "id": r.id,
                "name": r.name,
                "project_id": r.project_id,
                "status": r.status,
                "total_cases": r.total_cases,
                "passed_cases": r.passed_cases,
                "failed_cases": r.failed_cases,
                "duration_seconds": r.duration_seconds,
                "trigger_mode": r.trigger_mode,
                "created_at": r.created_at.isoformat() if r.created_at else "",
            }
            for r in recent_runs
        ],
    }


# ========== 安全设置 ==========
class URLWhitelistUpdate(BaseModel):
    urls: List[str]


@router.get("/settings/security")
def get_security_settings():
    """获取安全设置"""
    return {
        "url_whitelist": settings.URL_WHITELIST,
        "url_whitelist_enabled": settings.URL_WHITELIST_ENABLED,
        "tool_whitelist_enabled": settings.TOOL_WHITELIST_ENABLED,
        "readonly_mode": settings.READONLY_MODE,
        "auto_rollback_enabled": settings.AUTO_ROLLBACK_ENABLED,
        "default_tool_allowed": sorted(ToolWhitelist.DEFAULT_ALLOWED),
        "default_tool_blocked": sorted(ToolWhitelist.DEFAULT_BLOCKED),
    }


@router.put("/settings/security/url-whitelist")
def update_url_whitelist(data: URLWhitelistUpdate):
    """更新 URL 白名单"""
    settings.URL_WHITELIST = data.urls
    return {"message": "URL 白名单已更新", "urls": data.urls}


@router.put("/settings/security/url-whitelist-toggle")
def update_url_whitelist_enabled(enabled: bool = Query(...)):
    """更新 URL 白名单开关"""
    settings.URL_WHITELIST_ENABLED = enabled
    return {"message": f"URL白名单已{'开启' if enabled else '关闭'}"}


@router.put("/settings/security/readonly")
def update_readonly_mode(enabled: bool = Query(...)):
    """更新只读模式"""
    settings.READONLY_MODE = enabled
    return {"message": f"只读模式已{'开启' if enabled else '关闭'}"}


@router.put("/settings/security/auto-rollback")
def update_auto_rollback(enabled: bool = Query(...)):
    """更新自动回滚"""
    settings.AUTO_ROLLBACK_ENABLED = enabled
    return {"message": f"自动回滚已{'开启' if enabled else '关闭'}"}


@router.get("/settings/security/logs")
def get_security_logs(
    event_type: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db),
):
    """获取安全日志"""
    query = db.query(SecurityLog)
    if event_type:
        query = query.filter(SecurityLog.event_type == event_type)
    query = query.order_by(SecurityLog.created_at.desc())

    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "items": [
            {
                "id": log.id,
                "event_type": log.event_type,
                "target": log.target,
                "action": log.action,
                "result": log.result,
                "detail": log.detail[:300],
                "created_at": log.created_at.isoformat() if log.created_at else "",
            }
            for log in items
        ],
    }


# ========== 验证码/登录 ==========
class CookiesInput(BaseModel):
    cookies: str

class AutoLoginInput(BaseModel):
    url: str
    username_selector: str
    password_selector: str
    login_button_selector: str
    username: str
    password: str

@router.post("/captcha/login/{project_id}")
async def handle_captcha_login(project_id: int, db: Session = Depends(get_db)):
    """处理滑块验证码登录"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "项目不存在", 404)

    if not project.deploy_url:
        raise error_response(ErrorCode.MISSING_FIELD, "项目没有配置部署URL", 400)

    handler = CaptchaHandler()

    # Docker 环境（无显示器）返回 URL，提示用户手动操作
    if 'DISPLAY' not in os.environ:
        return {
            "message": "请在浏览器中打开地址完成登录，然后「粘贴 Cookie」",
            "login_url": project.deploy_url,
            "has_login": False,
            "manual_mode": True,
        }

    result = await handler.handle_slider_captcha(
        project_id=project_id,
        url=project.deploy_url,
        headless=False,
    )
    if result:
        return {"message": "登录成功，会话信息已保存", "has_login": True}
    return {"message": "登录超时或失败", "has_login": False}


@router.post("/captcha/cookies/{project_id}")
async def save_captcha_cookies(project_id: int, data: CookiesInput, db: Session = Depends(get_db)):
    """手动粘贴 Cookie 保存登录态"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "项目不存在", 404)

    import json
    try:
        cookies_list = json.loads(data.cookies)
        if not isinstance(cookies_list, list):
            raise ValueError("不是数组格式")
    except Exception as e:
        raise error_response(ErrorCode.INVALID_FORMAT, f"Cookie 格式错误: {e}", 400)

    record = db.query(LoginRecord).filter(
        LoginRecord.project_id == project_id,
    ).first()

    if not record:
        record = LoginRecord(
            project_id=project_id,
            url=project.deploy_url or "",
            cookies_data=data.cookies,
            session_valid=True,
            last_login_at=datetime.now(timezone.utc),
        )
        db.add(record)
    else:
        record.cookies_data = data.cookies
        record.session_valid = True
        record.last_login_at = datetime.now(timezone.utc)

    db.commit()
    logger.info(f"Cookie 已手动保存，项目ID: {project_id}")
    return {"message": f"Cookie 保存成功（共 {len(cookies_list)} 条）", "has_login": True}


@router.get("/captcha/status/{project_id}")
async def get_captcha_status(project_id: int, db: Session = Depends(get_db)):
    """获取验证码登录状态"""
    handler = CaptchaHandler()
    info = await handler.get_login_info(project_id)
    return info or {"has_login": False, "session_valid": False}


@router.delete("/captcha/login/{project_id}")
async def clear_captcha_login(project_id: int, db: Session = Depends(get_db)):
    """清除登录信息"""
    handler = CaptchaHandler()
    await handler.clear_login(project_id)
    return {"message": "登录信息已清除"}


@router.post("/captcha/check_session/{project_id}")
async def check_captcha_session(project_id: int, db: Session = Depends(get_db)):
    """手动触发登录会话有效性检查"""
    handler = CaptchaHandler()
    is_valid = await handler.check_session_validity(project_id)
    if is_valid:
        return {"message": "登录态有效", "session_valid": True}
    else:
        return {"message": "登录态已失效，请重新登录", "session_valid": False}


@router.post("/captcha/refresh_login/{project_id}")
async def refresh_captcha_login(project_id: int, db: Session = Depends(get_db)):
    """手动触发重新登录流程（弹出浏览器）"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        from fastapi import HTTPException
        raise HTTPException(status_code=404, detail="项目不存在")
    
    handler = CaptchaHandler()
    # 弹出浏览器让用户手动登录
    result = await handler.handle_slider_captcha(
        project_id=project_id,
        url=project.deploy_url,
        headless=False  # 必须非无头模式以便用户操作
    )
    
    if result:
        return {"message": "重新登录成功", "has_login": True}
    return {"message": "重新登录失败或超时", "has_login": False}


@router.post("/captcha/auto_login/{project_id}")
async def handle_auto_login(project_id: int, data: AutoLoginInput, db: Session = Depends(get_db)):
    """配置并尝试自动登录"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="项目不存在")

    from backend.security.encryption import encrypt_data
    encrypted_password = encrypt_data(data.password)

    handler = CaptchaHandler()
    result = await handler.auto_login(
        project_id=project_id,
        url=data.url,
        username_selector=data.username_selector,
        password_selector=data.password_selector,
        login_button_selector=data.login_button_selector,
        username=data.username,
        password=encrypted_password, # 传递加密后的密码
        headless=True # 自动登录通常在后台进行，使用无头模式
    )

    if result:
        return {"message": "自动登录配置成功并已尝试登录", "has_login": True}
    return {"message": "自动登录配置失败或登录失败", "has_login": False}


@router.get("/captcha/auto_login_config/{project_id}")
async def get_auto_login_config(project_id: int, db: Session = Depends(get_db)):
    """获取自动登录配置"""
    record = db.query(LoginRecord).filter(LoginRecord.project_id == project_id).first()
    if not record:
        return {"configured": False}
    
    return {
        "configured": True,
        "url": record.url,
        "username_selector": record.username_selector,
        "password_selector": record.password_selector,
        "login_button_selector": record.login_button_selector,
        "username": record.username,
        # 不返回密码
    }


# ========== 测试用例 ==========
@router.get("/test-cases")
def list_test_cases(
    project_id: Optional[int] = None,
    category: Optional[str] = None,
    db: Session = Depends(get_db),
):
    """获取测试用例列表"""
    query = db.query(TestCase)
    if project_id:
        query = query.filter(TestCase.project_id == project_id)
    if category:
        query = query.filter(TestCase.category == category)
    query = query.order_by(TestCase.created_at.desc())

    return {
        "items": [
            {
                "id": c.id,
                "project_id": c.project_id,
                "name": c.name,
                "description": c.description[:200] if c.description else "",
                "category": c.category,
                "priority": c.priority,
                "source": c.source,
                "tags": c.tags or [],
                "created_at": c.created_at.isoformat() if c.created_at else "",
            }
            for c in query.all()
        ],
    }


@router.delete("/test-cases/{case_id}")
def delete_test_case(case_id: int, db: Session = Depends(get_db)):
    """删除测试用例"""
    tc = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not tc:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试用例不存在", 404)
    db.delete(tc)
    db.commit()
    return {"message": "测试用例已删除"}


@router.get("/test-cases/{case_id}/code")
def get_test_case_code(case_id: int, db: Session = Depends(get_db)):
    """获取测试用例代码"""
    tc = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not tc:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试用例不存在", 404)
    return {"id": tc.id, "name": tc.name, "code": tc.code}


@router.put("/test-cases/{case_id}")
def update_test_case(case_id: int, data: dict, db: Session = Depends(get_db)):
    """更新测试用例名称或代码"""
    tc = db.query(TestCase).filter(TestCase.id == case_id).first()
    if not tc:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试用例不存在", 404)

    code = data.get("code", "").strip()
    name = data.get("name", "").strip()

    if code:
        from backend.security.code_validator import validate_test_code
        valid, error = validate_test_code(code)
        if not valid:
            raise error_response(ErrorCode.VALIDATION_ERROR, f"代码语法错误: {error}", 400)
        tc.code = code
    if name:
        tc.name = name

    db.commit()
    db.refresh(tc)
    return {"id": tc.id, "name": tc.name, "message": "测试用例已更新"}


@router.get("/test-cases/export/excel")
def export_test_cases_excel(
    project_id: Optional[int] = Query(None),
    db: Session = Depends(get_db),
):
    """导出测试用例为 Excel 文件"""
    query = db.query(TestCase)
    if project_id:
        query = query.filter(TestCase.project_id == project_id)
    query = query.order_by(TestCase.id.asc())
    cases = query.all()

    if not cases:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "没有可导出的测试用例", 404)

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_alignment = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # 表头
    headers = ['序号', '测试用例名称', '测试描述', '类别', '优先级', '操作步骤/代码']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # 数据行
    for idx, tc in enumerate(cases, 1):
        row_data = [
            idx,
            tc.name or '',
            tc.description or '',
            tc.category or '',
            tc.priority or '',
            tc.code or '',
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=idx + 1, column=col, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            cell.font = Font(name='微软雅黑', size=10)

    # 列宽
    ws.column_dimensions['A'].width = 6    # 序号
    ws.column_dimensions['B'].width = 30   # 名称
    ws.column_dimensions['C'].width = 40   # 描述
    ws.column_dimensions['D'].width = 12   # 类别
    ws.column_dimensions['E'].width = 10   # 优先级
    ws.column_dimensions['F'].width = 60   # 代码

    # 冻结首行
    ws.freeze_panes = 'A2'

    # 导出
    from fastapi.responses import Response
    import io
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f'attachment; filename="test_cases_project_{project_id or "all"}.xlsx"',
        },
    )


# ========== 系统信息 ==========
@router.get("/system/info")
def get_system_info():
    """获取系统信息"""
    import platform
    return {
        "app_name": settings.APP_NAME,
        "version": settings.APP_VERSION,
        "python_version": platform.python_version(),
        "platform": platform.platform(),
    }
