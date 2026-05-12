"""测试运行 API - 生成、执行测试用例并获取结果"""
import asyncio
import json
import os
import urllib.parse
from typing import Optional
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, WebSocket, WebSocketDisconnect
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse
from sqlalchemy.orm import Session
from pydantic import BaseModel, Field
from loguru import logger

from backend.database import get_db, Project, TestRun, TestCase, TestResult, AIConfig, SessionLocal
from backend.test_engine import TestGenerator, TestRunner, TestReporter
from backend.captcha import CaptchaHandler
from backend.security.url_validator import URLValidator
from backend.config import settings
from backend.errors import ErrorCode, error_response

router = APIRouter(prefix="/api/v1/test-runs", tags=["测试运行"])


class GenerateTestCaseRequest(BaseModel):
    """AI生成测试用例请求"""
    ai_config_id: int
    test_type: str = "功能测试"
    source_code_path: str = ""
    additional_context: str = ""


class RunTestRequest(BaseModel):
    """执行测试请求"""
    test_case_ids: list[int] = Field(default_factory=list)
    deploy_url: str = ""
    use_login: bool = True
    readonly_mode: bool = True
    headless: bool = True


# ---------- WebSocket 测试进度 ----------
connected_websockets = {}


@router.websocket("/ws/{run_id}")
async def test_progress_websocket(websocket: WebSocket, run_id: int):
    """测试进度 WebSocket"""
    await websocket.accept()
    connected_websockets[run_id] = websocket
    try:
        while True:
            await websocket.receive_text()  # 保持连接
    except WebSocketDisconnect:
        connected_websockets.pop(run_id, None)
    except Exception:
        connected_websockets.pop(run_id, None)


async def push_progress(data: dict):
    """推送进度到 WebSocket"""
    run_id = data.get("run_id")
    ws = connected_websockets.get(run_id)
    if ws:
        try:
            await ws.send_json(data)
        except Exception:
            connected_websockets.pop(run_id, None)


# ---------- API ----------
@router.get("")
def list_test_runs(
    project_id: Optional[int] = None,
    status: Optional[str] = None,
    page: int = Query(1, ge=1),
    page_size: int = Query(20, ge=1, le=100),
    db: Session = Depends(get_db),
):
    """获取测试运行列表"""
    query = db.query(TestRun)
    if project_id:
        query = query.filter(TestRun.project_id == project_id)
    if status:
        query = query.filter(TestRun.status == status)
    query = query.order_by(TestRun.created_at.desc())

    total = query.count()
    items = query.offset((page - 1) * page_size).limit(page_size).all()

    return {
        "total": total,
        "page": page,
        "page_size": page_size,
        "items": [
            {
                "id": r.id,
                "project_id": r.project_id,
                "name": r.name,
                "status": r.status,
                "trigger_mode": r.trigger_mode,
                "total_cases": r.total_cases,
                "passed_cases": r.passed_cases,
                "failed_cases": r.failed_cases,
                "error_cases": r.error_cases,
                "duration_seconds": r.duration_seconds,
                "created_at": r.created_at.isoformat() if r.created_at else "",
                "started_at": r.started_at.isoformat() if r.started_at else "",
                "completed_at": r.completed_at.isoformat() if r.completed_at else "",
            }
            for r in items
        ],
    }


@router.post("")
def create_test_run(
    project_id: int = Query(..., description="项目ID"),
    name: str = "自动化测试",
    db: Session = Depends(get_db),
):
    """创建测试运行"""
    project = db.query(Project).filter(Project.id == project_id).first()
    if not project:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "项目不存在", 404)

    run = TestRun(
        project_id=project_id,
        name=name,
        status="pending",
        trigger_mode="manual",
        total_cases=0,
    )
    db.add(run)
    db.commit()
    db.refresh(run)

    return {
        "id": run.id,
        "name": run.name,
        "status": run.status,
        "message": "测试运行已创建",
    }


@router.get("/{run_id}")
def get_test_run(run_id: int, db: Session = Depends(get_db)):
    """获取测试运行详情"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    results = db.query(TestResult).filter(TestResult.test_run_id == run_id).all()

    return {
        "run": {
            "id": run.id,
            "project_id": run.project_id,
            "name": run.name,
            "status": run.status,
            "trigger_mode": run.trigger_mode,
            "total_cases": run.total_cases,
            "passed_cases": run.passed_cases,
            "failed_cases": run.failed_cases,
            "error_cases": run.error_cases,
            "duration_seconds": run.duration_seconds,
            "readonly_mode": run.readonly_mode,
            "auto_rollback": run.auto_rollback,
            "report_path": run.report_path or "",
            "created_at": run.created_at.isoformat() if run.created_at else "",
            "started_at": run.started_at.isoformat() if run.started_at else "",
            "completed_at": run.completed_at.isoformat() if run.completed_at else "",
        },
        "results": [
            {
                "id": r.id,
                "test_case_id": r.test_case_id,
                "name": r.name,
                "func_name": f"test_{r.test_case_id}",
                "status": r.status,
                "duration_seconds": r.duration_seconds,
                "error_message": r.error_message[:500] if r.error_message else "",
                "screenshot_path": r.screenshot_path or "",
                "log_text": r.log_text[:1000] if r.log_text else "",
            }
            for r in results
        ],
    }


@router.post("/{run_id}/generate")
async def generate_test_cases(
    run_id: int,
    req: GenerateTestCaseRequest,
    db: Session = Depends(get_db),
):
    """AI 生成测试用例（后台异步执行）"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    ai_config = db.query(AIConfig).filter(AIConfig.id == req.ai_config_id).first()
    if not ai_config:
        raise error_response(ErrorCode.AI_CONFIG_INVALID, "AI 配置不存在或未激活", 400)

    project = db.query(Project).filter(Project.id == run.project_id).first()
    source_path = req.source_code_path or (project.source_code_path if project else "")

    if not source_path:
        raise error_response(ErrorCode.MISSING_FIELD, "请提供源代码路径", 400)

    # 更新状态为生成中
    run.status = "generating"
    run.name = f"AI生成测试 - {req.test_type}"
    db.commit()
    run_id_val = run.id

    # 后台异步执行 AI 生成（不阻塞前端）
    asyncio.create_task(_background_generate(
        run_id=run_id_val,
        ai_config={
            "provider": ai_config.provider,
            "api_key": ai_config.api_key,
            "api_base_url": ai_config.api_base_url,
            "model": ai_config.model,
            "temperature": ai_config.temperature,
            "max_tokens": ai_config.max_tokens,
        },
        project_id=run.project_id,
        source_path=source_path,
        test_type=req.test_type,
        additional_context=req.additional_context,
    ))

    return {
        "message": "AI 生成已启动，正在后台执行...",
        "run_id": run_id_val,
        "background": True,
    }


async def _background_generate(run_id: int, ai_config: dict, project_id: int,
                                source_path: str, test_type: str, additional_context: str):
    """后台执行 AI 测试用例生成"""
    try:
        # 查询数据库 Prompt 模板（优先 provider-specific，回退到通用模板）
        from backend.database import PromptTemplate
        template_body = None
        tdb = SessionLocal()
        try:
            template = tdb.query(PromptTemplate).filter(
                PromptTemplate.project_id == project_id,
                PromptTemplate.provider.in_(["", ai_config.get("provider", "")]),
                PromptTemplate.task_type == "test_generation",
            ).order_by(PromptTemplate.provider.desc()).first()
            template_body = template.template_body if template else None
        finally:
            tdb.close()

        generator = TestGenerator(ai_config, template_override=template_body)
        test_cases = await generator.generate_test_cases(
            project_id=project_id,
            source_path=source_path,
            test_type=test_type,
            additional_context=additional_context,
        )

        db = SessionLocal()
        try:
            run = db.query(TestRun).filter(TestRun.id == run_id).first()
            if run:
                run.total_cases = len(test_cases)
                run.status = "pending"
                db.commit()
            logger.info(f"后台生成完成: run_id={run_id}, count={len(test_cases)}")
            await push_progress({
                "type": "generate_complete",
                "run_id": run_id,
                "status": "completed",
                "progress": 100,
                "message": f"成功生成 {len(test_cases)} 个测试用例",
            })
        except Exception as e:
            logger.error(f"后台生成保存失败: {e}")
        finally:
            db.close()
    except ValueError as e:
        logger.error(f"代码验证失败: {e}")
        db = SessionLocal()
        try:
            run = db.query(TestRun).filter(TestRun.id == run_id).first()
            if run:
                run.status = "error"
                db.commit()
        finally:
            db.close()
        await push_progress({
            "type": "generate_error",
            "run_id": run_id,
            "status": "error",
            "progress": 0,
            "message": f"代码验证失败: {str(e)}",
        })
    except Exception as e:
        logger.error(f"后台生成失败: {e}")
        db = SessionLocal()
        try:
            run = db.query(TestRun).filter(TestRun.id == run_id).first()
            if run:
                run.status = "error"
                db.commit()
        finally:
            db.close()
        await push_progress({
            "type": "generate_error",
            "run_id": run_id,
            "status": "error",
            "progress": 0,
            "message": f"AI 生成失败: {str(e)}",
        })


@router.post("/{run_id}/execute")
async def execute_test_run(
    run_id: int,
    req: RunTestRequest,
    db: Session = Depends(get_db),
):
    """执行测试用例（异步后台执行 — 立即返回 202）"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    project = db.query(Project).filter(Project.id == run.project_id).first()
    deploy_url = req.deploy_url or (project.deploy_url if project else "")

    if not deploy_url:
        raise error_response(ErrorCode.MISSING_FIELD, "请提供部署URL", 400)

    # URL 安全校验（仅在白名单启用时校验）
    if settings.URL_WHITELIST_ENABLED:
        validator = URLValidator()
        if not validator.validate(deploy_url, whitelist=settings.URL_WHITELIST):
            raise error_response(ErrorCode.URL_NOT_WHITELISTED, f"URL 校验失败: {deploy_url} 不在白名单中", 400)

    # 获取测试用例
    if req.test_case_ids:
        cases = db.query(TestCase).filter(
            TestCase.id.in_(req.test_case_ids),
            TestCase.project_id == run.project_id,
        ).all()
    else:
        cases = db.query(TestCase).filter(
            TestCase.project_id == run.project_id,
        ).all()

    if not cases:
        raise error_response(ErrorCode.TEST_NO_CASES, "没有可执行的测试用例，请先生成", 400)

    # 获取登录信息
    login_info = None
    if req.use_login:
        captcha = CaptchaHandler()
        # 在执行前先校验登录态有效性
        is_valid = await captcha.check_session_validity(run.project_id)
        if not is_valid:
            logger.warning(f"项目 {run.project_id} 登录态校验失败，尝试获取已保存信息")
        
        login_info = await captcha.get_login_info(run.project_id)

    test_cases_data = [
        {"id": c.id, "name": c.name, "code": c.code}
        for c in cases
    ]

    # 启动后台执行（非阻塞）
    asyncio.create_task(_background_execute(
        run_id=run_id,
        test_cases=test_cases_data,
        deploy_url=deploy_url,
        login_info=login_info,
        readonly_mode=req.readonly_mode,
        headless=req.headless,
    ))

    return JSONResponse(
        status_code=202,
        content={"message": "测试执行已启动，正在后台运行...", "run_id": run_id, "background": True},
    )


async def _background_execute(run_id: int, test_cases: list, deploy_url: str,
                               login_info: dict, readonly_mode: bool, headless: bool):
    """后台任务: 通过 pytest 子进程执行测试"""
    runner = TestRunner(progress_callback=push_progress)
    try:
        result = await runner.run_test_cases(
            run_id=run_id,
            test_cases=test_cases,
            deploy_url=deploy_url,
            login_info=login_info,
            readonly_mode=readonly_mode,
            headless=headless,
        )
        TestReporter.generate_html_report(run_id)
        await push_progress({
            "type": "execute_complete",
            "run_id": run_id,
            "status": result["status"],
            "progress": 100,
            "message": f"测试完成: 通过{result['passed']}, 失败{result['failed']}, 错误{result['errors']}",
        })
    except Exception as e:
        logger.error(f"后台执行失败: {e}")
        db = SessionLocal()
        try:
            run = db.query(TestRun).filter(TestRun.id == run_id).first()
            if run:
                run.status = "error"
                run.completed_at = datetime.now(timezone.utc)
                db.commit()
        finally:
            db.close()
        await push_progress({
            "type": "execute_error",
            "run_id": run_id,
            "status": "error",
            "progress": 0,
            "message": f"测试执行失败: {str(e)}",
        })


@router.post("/{run_id}/retry/{test_case_id}")
async def retry_test_case(
    run_id: int,
    test_case_id: int,
    deploy_url: str = Query(..., description="部署URL"),
    db: Session = Depends(get_db),
):
    """重试单个失败的测试用例（隔离子进程执行）"""
    tc = db.query(TestCase).filter(TestCase.id == test_case_id).first()
    if not tc:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试用例不存在", 404)

    runner = TestRunner()
    result = await runner.retry_single_case(
        run_id=run_id,
        test_case_id=test_case_id,
        test_code=tc.code,
        test_name=tc.name,
        deploy_url=deploy_url,
        headless=True,
    )
    return {"result": result}


@router.post("/{run_id}/cancel")
def cancel_test_run(run_id: int, db: Session = Depends(get_db)):
    """取消测试运行"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    run.status = "cancelled"
    run.completed_at = datetime.now(timezone.utc)
    db.commit()

    # 通知正在运行的测试
    from backend.test_engine.runner import _active_runners
    runner = _active_runners.get(run_id)
    if runner:
        runner.cancel()

    return {"message": "测试已取消"}


@router.delete("/{run_id}")
def delete_test_run(run_id: int, db: Session = Depends(get_db)):
    """删除测试运行记录及所有结果"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)
    db.delete(run)
    db.commit()
    return {"message": "测试运行已删除"}


@router.get("/{run_id}/report")
def get_test_report(run_id: int, db: Session = Depends(get_db)):
    """获取测试报告"""
    report_data = TestReporter.get_report_data(run_id)
    if not report_data:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "报告不存在", 404)
    return report_data


@router.get("/{run_id}/export-excel")
def export_test_run_excel(run_id: int, db: Session = Depends(get_db)):
    """导出测试运行结果为 Excel"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    results = db.query(TestResult).filter(
        TestResult.test_run_id == run_id
    ).all()

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    import io

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试结果"

    # 样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='0F1923', end_color='0F1923', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell_align = Alignment(vertical='top', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='CCCCCC'),
        right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),
        bottom=Side(style='thin', color='CCCCCC'),
    )

    # 运行概要
    ws.cell(row=1, column=1, value="运行名称").font = header_font
    ws.cell(row=1, column=2, value=run.name or "").font = Font(name='微软雅黑', size=10)
    ws.cell(row=2, column=1, value="运行状态").font = header_font
    ws.cell(row=2, column=2, value=run.status).font = Font(name='微软雅黑', size=10)
    ws.cell(row=3, column=1, value="总用例").font = header_font
    ws.cell(row=3, column=2, value=run.total_cases or 0).font = Font(name='微软雅黑', size=10)
    ws.cell(row=4, column=1, value="通过").font = header_font
    ws.cell(row=4, column=2, value=run.passed_cases or 0).font = Font(name='微软雅黑', size=10)
    ws.cell(row=5, column=1, value="失败").font = header_font
    ws.cell(row=5, column=2, value=run.failed_cases or 0).font = Font(name='微软雅黑', size=10)
    ws.cell(row=6, column=1, value="错误").font = header_font
    ws.cell(row=6, column=2, value=run.error_cases or 0).font = Font(name='微软雅黑', size=10)
    ws.cell(row=7, column=1, value="耗时(秒)").font = header_font
    ws.cell(row=7, column=2, value=run.duration_seconds or 0).font = Font(name='微软雅黑', size=10)

    # 空行
    start_row = 9

    # 结果表头
    headers = ['序号', '用例名称', '状态', '耗时(秒)', '错误信息', '执行日志']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 结果数据
    for idx, r in enumerate(results, 1):
        row_data = [
            idx,
            r.name or '',
            r.status or '',
            r.duration_seconds or 0,
            (r.error_message or '')[:500],
            (r.log_text or '')[:500],
        ]
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=start_row + idx, column=col, value=val)
            cell.alignment = cell_align
            cell.border = thin_border
            cell.font = Font(name='微软雅黑', size=10)

    # 列宽
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 50

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    from fastapi.responses import Response
    safe_name = (run.name or f"report_{run_id}")
    # HTTP Header只能ASCII，中文用RFC5987编码
    ascii_filename = f"test_result_{run_id}.xlsx"
    encoded_filename = urllib.parse.quote(safe_name + ".xlsx", safe='')
    
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/{run_id}/report/html")
def get_test_report_html(run_id: int, db: Session = Depends(get_db)):
    """获取 HTML 测试报告"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    # 优先返回数据库中的 HTML
    if run.report_html:
        return HTMLResponse(content=run.report_html, media_type="text/html; charset=utf-8")

    # 其次尝试从文件读取
    report_path = run.report_path
    if report_path and os.path.exists(report_path):
        with open(report_path, "r", encoding="utf-8") as f:
            html_content = f.read()
        return HTMLResponse(content=html_content, media_type="text/html; charset=utf-8")

    raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "HTML 报告不存在，请先执行测试", 404)


@router.get("/{run_id}/screenshot/{result_id}")
def get_test_screenshot(run_id: int, result_id: int, db: Session = Depends(get_db)):
    """获取测试截图"""
    result = db.query(TestResult).filter(
        TestResult.id == result_id,
        TestResult.test_run_id == run_id,
    ).first()
    if not result or not result.screenshot_path:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "截图不存在", 404)

    screenshot_path = result.screenshot_path
    if os.path.exists(screenshot_path):
        return FileResponse(screenshot_path, media_type="image/png")

    # 尝试相对路径
    abs_path = os.path.join(settings.REPORT_DIR, f"run_{run_id}", os.path.basename(screenshot_path))
    if os.path.exists(abs_path):
        return FileResponse(abs_path, media_type="image/png")

    raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "截图文件不存在", 404)


@router.get("/{run_id}/export-report")
def export_test_report(run_id: int, db: Session = Depends(get_db)):
    """导出 HTML 测试报告（下载）"""
    run = db.query(TestRun).filter(TestRun.id == run_id).first()
    if not run:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "测试运行不存在", 404)

    html_content = None
    if run.report_html:
        html_content = run.report_html
    elif run.report_path and os.path.exists(run.report_path):
        with open(run.report_path, "r", encoding="utf-8") as f:
            html_content = f.read()

    if not html_content:
        raise error_response(ErrorCode.RESOURCE_NOT_FOUND, "报告不存在，请先执行测试", 404)

    import urllib.parse
    safe_name = run.name.replace(" ", "_").replace("/", "_") or f"report_{run_id}"
    # HTTP Header 必须 ASCII，中文文件名用 filename*（RFC 5987）编码
    ascii_filename = f"report_{run_id}.html"
    encoded_filename = urllib.parse.quote(f"{safe_name}.html", safe='')
    from fastapi.responses import Response
    return Response(
        content=html_content.encode("utf-8"),
        media_type="text/html; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename=\"{ascii_filename}\"; filename*=UTF-8''{encoded_filename}",
        },
    )
